import googlemaps

from config import API_KEY

from openpyxl import load_workbook, Workbook
import json
from collections import defaultdict
import math
from time import sleep


def get_ICAO_codes():
    wb = load_workbook("Airport_Codes.xlsx")
    last_row = 2922
    sheet = wb["Vendors"]
    codes = []
    for i in range(2, last_row + 1):
        icao_codes = sheet["G{}".format(i)].value
        iata_codes = sheet["H{}".format(i)].value
        if icao_codes and not iata_codes and icao_codes != "N/A":
            codes.append(icao_codes)
    return codes


def get_ICAO_codes_mapped(codes):
    code_to_coord = {}
    wb = load_workbook("Airport Master.xlsx")
    last_row = 2728
    sheet = wb["US Airports Master"]
    for i in range(2, last_row + 1):
        airport_code = sheet["M{}".format(i)].value
        lat, long = sheet["E{}".format(i)].value, sheet["F{}".format(i)].value
        if airport_code in codes:
            code_to_coord[airport_code] = (lat, long)
    return code_to_coord


def get_IATA_codes():
    wb = load_workbook("Airport_Codes.xlsx")
    last_row = 2922
    sheet = wb["Vendors"]
    codes = []
    for i in range(2, last_row + 1):
        airport_code = sheet["H{}".format(i)].value
        if airport_code and airport_code != "N/A":
            codes.append(airport_code)
    return codes


def get_IATA_codes_mapped(codes):
    code_to_coord = {}
    wb = load_workbook("Airport Master.xlsx")
    last_row = 2728
    sheet = wb["US Airports Master"]
    for i in range(2, last_row + 1):
        airport_code = sheet["N{}".format(i)].value
        lat, long = sheet["E{}".format(i)].value, sheet["F{}".format(i)].value
        if airport_code in codes:
            code_to_coord[airport_code] = (lat, long)
    return code_to_coord


def get_IATA_businesses():
    iata_codes = get_IATA_codes()
    iata_code_to_coord = get_IATA_codes_mapped(iata_codes)

    results_dict = {}
    client = googlemaps.Client(key=API_KEY)

    for code in iata_codes:
        query = "aircraft maintenance"
        try:
            lat, lon = list(map(str, iata_code_to_coord[code]))
            print(query, code)
        except:
            print("NO DATA FOR AIRPORT CODE: {}".format(code))
            results_dict[code] = []
            continue

        results = client.places_nearby(keyword=query, location=(lat, lon), radius=48000)["results"]
        print(results)
        sleep(1)

        results_dict[code] = results
    with open('IATA_results.json', 'w') as outfile:
        json.dump(results_dict, outfile)


def get_ICAO_businesses():
    icao_codes = get_ICAO_codes()
    icao_code_to_coord = get_ICAO_codes_mapped(icao_codes)

    results_dict = {}
    client = googlemaps.Client(key=API_KEY)

    for code in icao_codes:
        query = "aircraft maintenance"
        try:
            lat, lon = list(map(str, icao_code_to_coord[code]))
            print(query, code)
        except:
            print("NO DATA FOR AIRPORT CODE: {}".format(code))
            results_dict[code] = []
            continue

        results = client.places_nearby(keyword=query, location=(lat, lon), radius=48000)["results"]
        print(results)
        sleep(1)

        results_dict[code] = results

    with open('ICAO_results.json', 'w') as outfile:
        json.dump(results_dict, outfile)


def check_IATA_dupes():
    with open("IATA_results.json", "r") as f:
        results = json.load(f)

    all_values = []
    for code in results:
        all_values.extend(results[code])
    place_ids = []
    for val in all_values:
        place_ids.append(val["place_id"])

    print(len(place_ids))
    print(len(set(place_ids)))
    unique_ids = list(set(place_ids))

    # map place id to the airports its near
    place_id_to_airport = defaultdict(list)
    for code in results:
        values = results[code]
        for value in values:
            if value['place_id'] in unique_ids:
                place_id_to_airport[value['place_id']].append(code)
            else:
                pass

    # map place id to its lat/long
    place_id_to_coord = {}
    for val in all_values:
        location = val["geometry"]["location"]
        lat, long = location["lat"], location["lng"]
        place_id_to_coord[val["place_id"]] = [lat, long]

    # map airport code to lat/long
    iata_codes = get_IATA_codes()
    iata_code_to_coord = get_IATA_codes_mapped(iata_codes)

    for place_id, airports_codes in place_id_to_airport.items():
        if len(airports_codes) == 1:
            continue
        place_id_coords = place_id_to_coord[place_id]
        closest_airport = min(airports_codes, key=lambda x: distance(iata_code_to_coord[x], place_id_coords))
        place_id_to_airport[place_id] = [closest_airport]

    new_results = {}
    for code in results:
        values = results[code]
        new_values = []
        for value in values:
            if code == place_id_to_airport[value['place_id']][0]:
                new_values.append(value)
            else:
                pass
        new_results[code] = new_values
    all_new_values = []
    for code in new_results:
        all_new_values.extend(new_results[code])
    print(len(all_new_values))
    with open('filtered_IATA_results.json', 'w') as outfile:
        json.dump(new_results, outfile)


def check_ICAO_dupes():
    with open("ICAO_results.json", "r") as f:
        results = json.load(f)

    all_values = []
    for code in results:
        all_values.extend(results[code])
    place_ids = []
    for val in all_values:
        place_ids.append(val["place_id"])

    print(len(place_ids))
    print(len(set(place_ids)))
    unique_ids = list(set(place_ids))

    # map place id to the airports its near
    place_id_to_airport = defaultdict(list)
    for code in results:
        values = results[code]
        for value in values:
            if value['place_id'] in unique_ids:
                place_id_to_airport[value['place_id']].append(code)
            else:
                pass

    # map place id to its lat/long
    place_id_to_coord = {}
    for val in all_values:
        location = val["geometry"]["location"]
        lat, long = location["lat"], location["lng"]
        place_id_to_coord[val["place_id"]] = [lat, long]

    # map airport code to lat/long
    icao_codes = get_ICAO_codes()
    icao_code_to_coord = get_ICAO_codes_mapped(icao_codes)

    for place_id, airports_codes in place_id_to_airport.items():
        if len(airports_codes) == 1:
            continue
        place_id_coords = place_id_to_coord[place_id]
        closest_airport = min(airports_codes, key=lambda x: distance(icao_code_to_coord[x], place_id_coords))
        place_id_to_airport[place_id] = [closest_airport]

    new_results = {}
    for code in results:
        values = results[code]
        new_values = []
        for value in values:
            if code == place_id_to_airport[value['place_id']][0]:
                new_values.append(value)
            else:
                pass
        new_results[code] = new_values
    all_new_values = []
    for code in new_results:
        all_new_values.extend(new_results[code])
    print(len(all_new_values))
    with open('filtered_ICAO_results.json', 'w') as outfile:
        json.dump(new_results, outfile)


def check_dupes():
    with open("IATA_results.json", "r") as f:
        IATA_results = json.load(f)

    with open("ICAO_results.json", "r") as f:
        ICAO_results = json.load(f)

    all_results = {}
    all_results.update(IATA_results)
    all_results.update(ICAO_results)

    all_values = []
    for code in all_results:
        all_values.extend(all_results[code])
    place_ids = []
    for val in all_values:
        place_ids.append(val["place_id"])

    print(len(place_ids))
    print(len(set(place_ids)))
    unique_ids = list(set(place_ids))

    # map place id to the airports its near
    place_id_to_airport = defaultdict(list)
    for code in all_results:
        values = all_results[code]
        for value in values:
            # if value['place_id'] in unique_ids:
                place_id_to_airport[value['place_id']].append(code)
            # else:
            #     pass


    # map place id to its lat/long
    place_id_to_coord = {}
    for val in all_values:
        location = val["geometry"]["location"]
        lat, long = location["lat"], location["lng"]
        place_id_to_coord[val["place_id"]] = [lat, long]

    # map airport code to lat/long
    iata_codes = get_IATA_codes()
    iata_code_to_coord = get_IATA_codes_mapped(iata_codes)

    # map airport code to lat/long
    icao_codes = get_ICAO_codes()
    icao_code_to_coord = get_ICAO_codes_mapped(icao_codes)

    all_codes_to_coords = {}
    all_codes_to_coords.update(icao_code_to_coord)
    all_codes_to_coords.update(iata_code_to_coord)

    for place_id, airports_codes in place_id_to_airport.items():
        if len(airports_codes) == 1:
            continue
        place_id_coords = place_id_to_coord[place_id]
        closest_airport = min(airports_codes, key=lambda x: distance(all_codes_to_coords[x], place_id_coords))
        place_id_to_airport[place_id] = [closest_airport]

    new_results = {}
    for code in all_results:
        values = all_results[code]
        new_values = []
        for value in values:
            if code == place_id_to_airport[value['place_id']][0]:
                new_values.append(value)
            else:
                pass
        new_results[code] = new_values
    all_new_values = []
    for code in new_results:
        all_new_values.extend(new_results[code])
    print(len(all_new_values))
    with open('filtered_results.json', 'w') as outfile:
        json.dump(new_results, outfile)


def get_contact_info_IATA():
    client = googlemaps.Client(key=API_KEY)

    with open("filtered_IATA_results.json", "r") as f:
        result = json.load(f)

    # map airport codes to list of place_ids
    code_to_place_ids = {}
    for code in result:
        place_ids = []
        for place in result[code]:
            place_ids.append(place["place_id"])
        code_to_place_ids[code] = place_ids

    # make places detail request with place_id
    details = []
    for code in code_to_place_ids:
        place_ids = code_to_place_ids[code]

        # create list of tuples of Airport ID, Name, Address, Phone
        for place_id in place_ids:
            result = client.place(place_id=place_id)["result"]
            # print(result)
            business_name = result.get("name")
            address = result.get('formatted_address', "")
            phone = result.get('formatted_phone_number', "")

            place_details = [code, business_name, address, phone]
            details.append(place_details)
            print(place_details)
    workbook = Workbook()
    sheet = workbook.active

    sheet["A1"] = "IATA"
    sheet["B1"] = "Name"
    sheet["C1"] = "Address"
    sheet["D1"] = "Phone"

    for i in range(2, len(details) + 2):
        for letter, detail in zip(["A", "B", "C", "D"], details[i - 2]):
            sheet["{}{}".format(letter, i)] = detail

    # save list as excel
    workbook.save(filename="Place_Details_IATA.xlsx")


def get_contact_info_ICAO():
    client = googlemaps.Client(key=API_KEY)

    with open("filtered_ICAO_results.json", "r") as f:
        result = json.load(f)

    # map airport codes to list of place_ids
    code_to_place_ids = {}
    for code in result:
        place_ids = []
        for place in result[code]:
            place_ids.append(place["place_id"])
        code_to_place_ids[code] = place_ids

    # make places detail request with place_id
    details = []
    for code in code_to_place_ids:
        place_ids = code_to_place_ids[code]

        # create list of tuples of Airport ID, Name, Address, Phone
        for place_id in place_ids:
            result = client.place(place_id=place_id)["result"]
            # print(result)
            business_name = result.get("name")
            address = result.get('formatted_address', "")
            phone = result.get('formatted_phone_number', "")

            place_details = [code, business_name, address, phone]
            details.append(place_details)
            print(place_details)
    workbook = Workbook()
    sheet = workbook.active

    sheet["A1"] = "ICAO"
    sheet["B1"] = "Name"
    sheet["C1"] = "Address"
    sheet["D1"] = "Phone"

    for i in range(2, len(details) + 2):
        for letter, detail in zip(["A", "B", "C", "D"], details[i - 2]):
            sheet["{}{}".format(letter, i)] = detail

    # save list as excel
    workbook.save(filename="Place_Details_ICAO.xlsx")


def get_contact_info():
    client = googlemaps.Client(key=API_KEY)
    iata_codes = get_IATA_codes()
    icao_codes = get_ICAO_codes()
    _dict = codes_to_states()

    with open("filtered_results.json", "r") as f:
        result = json.load(f)

    # map airport codes to list of place_ids
    code_to_place_ids = {}
    for code in result:
        place_ids = []
        for place in result[code]:
            place_ids.append(place["place_id"])
        code_to_place_ids[code] = place_ids

    # make places detail request with place_id
    details = []
    for code in code_to_place_ids:
        place_ids = code_to_place_ids[code]

        # create list of tuples of Airport ID, Name, Address, Phone
        for place_id in place_ids:
            result = client.place(place_id=place_id)["result"]
            # print(result)
            business_name = result.get("name")
            address = result.get('formatted_address', "")
            phone = result.get('formatted_phone_number', "")

            place_details = [code, business_name, address, phone]
            details.append(place_details)
            print(place_details)
            sleep(1)

    workbook = Workbook()
    sheet = workbook.active

    sheet["A1"] = "Airport Code"
    sheet["B1"] = "Name"
    sheet["C1"] = "Address"
    sheet["D1"] = "Phone"
    sheet["E1"] = "Code Type"
    sheet["F1"] = "State"

    for i in range(2, len(details) + 2):
        for letter, detail in zip(["A", "B", "C", "D"], details[i - 2]):
            sheet["{}{}".format(letter, i)] = detail
        code = details[i - 2][0]
        if code in icao_codes:
            sheet["E{}".format(i)] = "ICAO"
        elif code in iata_codes:
            sheet["E{}".format(i)] = "IATA"
        sheet["F{}".format(i)] = _dict[code]

    # save list as excel
    workbook.save(filename="Place_Details.xlsx")


def final_steps_ICAO():
    _dict = codes_to_states()
    wb = load_workbook("Place_Details_ICAO.xlsx")
    sheet = wb.active
    last_row = sheet.max_row
    for i in range(2, last_row + 1):
        iata_code = sheet["{}{}".format("A", i)].value
        icao_code = sheet["{}{}".format("B", i)].value
        if iata_code:
            sheet["F{}".format(i)] = _dict[iata_code]
        else:
            sheet["F{}".format(i)] = _dict[icao_code]

    wb.save(filename="Place_Details_ICAO.xlsx")


def final_steps_IATA():
    _dict = codes_to_states()
    wb = load_workbook("Place_Details_IATA.xlsx")
    sheet = wb.active
    last_row = sheet.max_row
    for i in range(2, last_row + 1):
        iata_code = sheet["{}{}".format("A", i)].value
        icao_code = sheet["{}{}".format("B", i)].value
        if iata_code:
            sheet["F{}".format(i)] = _dict[iata_code]
        else:
            sheet["F{}".format(i)] = _dict[icao_code]

    wb.save(filename="Place_Details_IATA.xlsx")


def distance(lat_long0, lat_long1):
    R = 6373.0
    lat1 = math.radians(lat_long0[0])
    lon1 = math.radians(lat_long0[1])
    lat2 = math.radians(lat_long1[0])
    lon2 = math.radians(lat_long1[1])
    dlon = lon2 - lon1
    dlat = lat2 - lat1
    a = math.sin(dlat / 2) ** 2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon / 2) ** 2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    distance = R * c

    return distance


def codes_to_states():
    icao_col = "M"
    iata_col = "N"
    state_col = "J"
    state_to_initials = {
        'Alabama': 'AL',
        'Alaska': 'AK',
        'American Samoa': 'AS',
        'Arizona': 'AZ',
        'Arkansas': 'AR',
        'California': 'CA',
        'Colorado': 'CO',
        'Connecticut': 'CT',
        'Delaware': 'DE',
        'District of Columbia': 'DC',
        'Florida': 'FL',
        'Georgia': 'GA',
        'Guam': 'GU',
        'Hawaii': 'HI',
        'Idaho': 'ID',
        'Illinois': 'IL',
        'Indiana': 'IN',
        'Iowa': 'IA',
        'Kansas': 'KS',
        'Kentucky': 'KY',
        'Louisiana': 'LA',
        'Maine': 'ME',
        'Maryland': 'MD',
        'Massachusetts': 'MA',
        'Michigan': 'MI',
        'Minnesota': 'MN',
        'Mississippi': 'MS',
        'Missouri': 'MO',
        'Montana': 'MT',
        'Nebraska': 'NE',
        'Nevada': 'NV',
        'New Hampshire': 'NH',
        'New Jersey': 'NJ',
        'New Mexico': 'NM',
        'New York': 'NY',
        'North Carolina': 'NC',
        'North Dakota': 'ND',
        'Northern Mariana Islands': 'MP',
        'Ohio': 'OH',
        'Oklahoma': 'OK',
        'Oregon': 'OR',
        'Pennsylvania': 'PA',
        'Puerto Rico': 'PR',
        'Rhode Island': 'RI',
        'South Carolina': 'SC',
        'South Dakota': 'SD',
        'Tennessee': 'TN',
        'Texas': 'TX',
        'Utah': 'UT',
        'Vermont': 'VT',
        'Virgin Islands': 'VI',
        'Virginia': 'VA',
        'Washington': 'WA',
        'West Virginia': 'WV',
        'Wisconsin': 'WI',
        'Wyoming': 'WY'
    }
    initials_to_state = {y: x for x, y in state_to_initials.items()}
    code_to_state = {}
    wb = load_workbook("Airport Master.xlsx")
    last_row = 2728
    sheet = wb["US Airports Master"]
    for i in range(2, last_row + 1):
        iata_code = sheet["{}{}".format(iata_col, i)].value
        icao_code = sheet["{}{}".format(icao_col, i)].value
        state = initials_to_state[sheet["{}{}".format(state_col, i)].value.split("-")[1]]
        code_to_state[iata_code] = state
        code_to_state[icao_code] = state
    return code_to_state


if __name__ == '__main__':
    # get_ICAO_businesses()
    # get_IATA_businesses()
    check_dupes()
    # get_contact_info()

